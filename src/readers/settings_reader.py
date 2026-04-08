from datetime import date, datetime
from typing import Any

import openpyxl

from ..config import (
    BASIC_PARAM_NAME_COL,
    BASIC_PARAM_RANGES,
    BASIC_PARAM_VALUE_COL,
    CURRENT_DATE_COL,
    CURRENT_DATE_ROW,
    DEVIATION_COLUMNS,
    DEVIATION_START_ROW,
    EXCEPTION_COLUMNS,
    EXCEPTION_COUNT_COL,
    EXCEPTION_COUNT_ROW,
    EXCEPTION_START_ROW,
    SHEET_BASIC_SETTINGS,
    SHEET_EXCEPTIONS,
)
from ..models.exception_rule import ExceptionRule
from ..models.settings import BasicSettings, DeviationCoeffRow
from ..validators import (
    BasicSettingsValidator,
    DeviationTableValidator,
    ExceptionRuleValidator,
)
from .base_reader import BaseReader

RawDict = dict[str, Any]


class SettingsReader(BaseReader):
    def __init__(self, filepath: str, strict: bool = True) -> None:
        super().__init__(filepath)
        self.strict = strict
        self._warnings: list[str] = []

    @property
    def warnings(self) -> list[str]:
        return list(self._warnings)

    def read(self) -> dict[str, Any]:
        self._warnings.clear()

        raw = self._read_all_raw()
        validated = self._validate_raw(raw)
        result = self._build_all(validated)
        self._validate_built(result)

        return result

    def _read_all_raw(self) -> dict[str, Any]:
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        try:
            ws_settings = wb[SHEET_BASIC_SETTINGS]
            ws_exceptions = wb[SHEET_EXCEPTIONS]
            return {
                "raw_params": self._read_raw_basic_settings(ws_settings),
                "current_date": self._read_current_date(ws_settings),
                "raw_deviation": self._read_raw_deviation_table(ws_settings),
                "raw_exceptions": self._read_raw_exceptions(ws_exceptions),
            }
        finally:
            wb.close()

    def _read_raw_basic_settings(self, ws: Any) -> RawDict:
        raw: RawDict = {}
        for start, end in BASIC_PARAM_RANGES:
            for row in range(start, end):
                name = str(ws.cell(row=row, column=BASIC_PARAM_NAME_COL).value).strip()
                raw[name] = ws.cell(row=row, column=BASIC_PARAM_VALUE_COL).value
        return raw

    def _read_current_date(self, ws: Any) -> date:
        raw = ws.cell(row=CURRENT_DATE_ROW, column=CURRENT_DATE_COL).value
        if isinstance(raw, datetime):
            return raw.date()
        if isinstance(raw, date):
            return raw
        raise ValueError(f"Не удалось прочитать current_date из F2: {raw!r}")

    def _read_raw_deviation_table(self, ws: Any) -> list[RawDict]:
        rows: list[RawDict] = []
        r = DEVIATION_START_ROW
        while True:
            label_raw = ws.cell(row=r, column=DEVIATION_COLUMNS["label"]).value
            if label_raw is None or str(label_raw).strip() == "":
                break
            row_data: RawDict = {"excel_row": r}
            for key, col in DEVIATION_COLUMNS.items():
                value = ws.cell(row=r, column=col).value
                row_data[key] = str(value).strip() if key == "label" else value
            rows.append(row_data)
            r += 1
        return rows

    def _read_raw_exceptions(self, ws: Any) -> list[RawDict]:
        count = int(
            ws.cell(row=EXCEPTION_COUNT_ROW, column=EXCEPTION_COUNT_COL).value or 0
        )
        rows: list[RawDict] = []
        for r in range(EXCEPTION_START_ROW, EXCEPTION_START_ROW + count):
            rows.append(
                {
                    key: ws.cell(row=r, column=col).value
                    for key, col in EXCEPTION_COLUMNS.items()
                }
            )
        return rows

    def _validate_raw(self, raw: dict[str, Any]) -> dict[str, Any]:
        errors: list[str] = []
        errors.extend(BasicSettingsValidator().validate_raw(raw["raw_params"]))
        errors.extend(DeviationTableValidator().validate_raw(raw["raw_deviation"]))
        errors.extend(ExceptionRuleValidator().validate_raw(raw["raw_exceptions"]))

        if errors:
            self._handle_errors(errors)

        return raw

    def _build_all(self, raw: dict[str, Any]) -> dict[str, Any]:
        return {
            "basic_settings": self._build_basic_settings(raw["raw_params"]),
            "current_date": raw["current_date"],
            "deviation_table": self._build_deviation_table(raw["raw_deviation"]),
            "exceptions": self._build_exceptions(raw["raw_exceptions"]),
        }

    def _build_basic_settings(self, raw: RawDict) -> BasicSettings:
        to = self._to_float
        return BasicSettings(
            base_rate=to(raw["Base_Rate"]),
            base_share=to(raw["Base_Share"]),
            min_rate=to(raw["Min_Rate"]),
            max_rate=to(raw["Max_Rate"]),
            min_share=to(raw["Min_Share"]),
            max_share=to(raw["Max_Share"]),
            rate_multiplier=to(raw["Rate_Multiplier"]),
            share_multiplier=to(raw["Share_Multiplier"]),
        )

    def _build_deviation_table(
        self, raw_rows: list[RawDict]
    ) -> list[DeviationCoeffRow]:
        to = self._to_float
        return [
            DeviationCoeffRow(
                load_label=str(row["label"]),
                min_dev=to(row["min_dev"]),
                max_dev=to(row["max_dev"]),
                coef_rate=to(row["coef_rate"]),
                coef_share=to(row["coef_share"]),
            )
            for row in raw_rows
        ]

    def _build_exceptions(self, raw_rows: list[RawDict]) -> list[ExceptionRule]:
        to_f = self._to_float
        to_d = self._to_date
        return [
            ExceptionRule(
                rule_id=int(row["rule_id"]),
                origin=str(row["origin"] or "").strip().upper(),
                destination=str(row["destination"] or "").strip().upper(),
                date_from=to_d(row["date_from"]),
                date_to=to_d(row["date_to"]),
                fixed_rate=to_f(row["fixed_rate"]),
                fixed_share=to_f(row["fixed_share"]),
                reason=str(row["reason"] or ""),
                priority=int(row["priority"]),
            )
            for row in raw_rows
        ]

    def _validate_built(self, result: dict[str, Any]) -> None:
        errors: list[str] = []
        errors.extend(BasicSettingsValidator().validate(result["basic_settings"]))
        errors.extend(DeviationTableValidator().validate(result["deviation_table"]))
        errors.extend(ExceptionRuleValidator().validate(result["exceptions"]))

        if errors:
            self._handle_errors(errors)

    def _handle_errors(self, errors: list[str]) -> None:
        if self.strict:
            raise ValueError(
                "Файл настроек содержит ошибки:\n"
                + "\n".join(f"  • {e}" for e in errors)
            )
        self._warnings.extend(errors)

    @staticmethod
    def _to_float(value: Any) -> float:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).replace(",", ".").replace("%", "").strip()
        try:
            return float(s)
        except ValueError:
            return 0.0

    @staticmethod
    def _to_date(value: Any) -> date:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        raise ValueError(f"Невозможно преобразовать в дату: {value!r}")
